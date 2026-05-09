from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from ev_flex_trading.reporting.excel_report import generate_excel_report


def test_excel_report_creates_required_workbook(tmp_path: Path) -> None:
    workbook_path = tmp_path / "report.xlsx"
    manifest_path = tmp_path / "manifest.csv"

    generated_path, manifest = generate_excel_report(
        output_path=workbook_path,
        manifest_path=manifest_path,
        auto_generate_inputs=True,
    )

    assert generated_path.exists()
    assert manifest_path.exists()
    assert not manifest.empty

    workbook = load_workbook(generated_path, data_only=False)
    expected_sheets = [
        "README",
        "Daily Summary",
        "Baseline vs Optimized",
        "Scheduled vs Actual",
        "Settlement Exposure",
        "Market Metrics",
        "Exceptions",
        "Fleet Requirements",
        "Market Prices",
        "Baseline Schedule",
        "Optimized Schedule",
        "Actual Charging",
        "Assumptions",
    ]
    assert workbook.sheetnames == expected_sheets


def test_excel_report_has_public_safe_disclaimer(tmp_path: Path) -> None:
    workbook_path = tmp_path / "report.xlsx"
    manifest_path = tmp_path / "manifest.csv"

    generate_excel_report(
        output_path=workbook_path,
        manifest_path=manifest_path,
        auto_generate_inputs=True,
    )

    workbook = load_workbook(workbook_path, data_only=True)
    readme_values = [cell.value for row in workbook["README"].iter_rows() for cell in row]
    assumptions_values = [cell.value for row in workbook["Assumptions"].iter_rows() for cell in row]
    combined = " ".join(str(value) for value in [*readme_values, *assumptions_values] if value)

    assert "not a production trading, dispatch, or settlement system" in combined
    assert "not official BSC settlement" in combined


def test_excel_report_key_summary_cells_populated(tmp_path: Path) -> None:
    workbook_path = tmp_path / "report.xlsx"
    manifest_path = tmp_path / "manifest.csv"

    generate_excel_report(
        output_path=workbook_path,
        manifest_path=manifest_path,
        auto_generate_inputs=True,
    )

    workbook = load_workbook(workbook_path, data_only=True)
    sheet = workbook["Daily Summary"]
    assert sheet["A1"].value == "Daily Summary"
    assert sheet["A5"].value is not None
    assert sheet["D5"].value is not None


def test_excel_report_manifest_is_written(tmp_path: Path) -> None:
    workbook_path = tmp_path / "report.xlsx"
    manifest_path = tmp_path / "manifest.csv"

    generate_excel_report(
        output_path=workbook_path,
        manifest_path=manifest_path,
        auto_generate_inputs=True,
    )

    manifest = pd.read_csv(manifest_path)
    assert "workbook" in set(manifest["input_name"])
